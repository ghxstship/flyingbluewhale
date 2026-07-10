#!/usr/bin/env python3
"""Build the ATLVS Master XPMS Catalog.

Combines the FrozenPhoenix Universal Advance Seed Catalog (351 items) with the
existing flyingbluewhale canonical master_catalog_items, dedups, and classifies
every line item from scratch into the XPMS taxonomy (Department · URID ·
Discipline · Tier · Phase · XYZ) using the canonical xpms_registry team codes.

Emits:
  supabase/migrations/20260608120000_xpms_master_catalog_schema.sql
  supabase/migrations/20260608120100_xpms_master_catalog_data.sql
  xpms-master-catalog.xlsx
"""
import re, json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

FP = "/Users/julianclarkson/Documents/FrozenPhoenix/supabase/migrations"
ROOT = "/Users/julianclarkson/Documents/flyingbluewhale"
ORG = "68672cc3-0667-4234-ad77-49325e173175"  # demo org

# ── SQL parse helpers ──────────────────────────────────────────────────────────
def split_top(s):
    out, buf, depth, instr = [], [], 0, False; i = 0
    while i < len(s):
        c = s[i]
        if instr:
            if c == "'":
                if i + 1 < len(s) and s[i + 1] == "'": buf.append("''"); i += 2; continue
                instr = False
            buf.append(c)
        else:
            if c == "'": instr = True; buf.append(c)
            elif c in "([": depth += 1; buf.append(c)
            elif c in ")]": depth -= 1; buf.append(c)
            elif c == "," and depth == 0: out.append("".join(buf).strip()); buf = []
            else: buf.append(c)
        i += 1
    if buf: out.append("".join(buf).strip())
    return out

def unq(v):
    v = v.strip()
    if v.upper() == "NULL": return None
    v = re.sub(r"::\w+(\s*\[\])?$", "", v).strip()
    if v.startswith("'") and v.endswith("'"): return v[1:-1].replace("''", "'")
    return v

def split_rows(body):
    depth, instr, start = 0, False, None; i = 0
    while i < len(body):
        c = body[i]
        if instr:
            if c == "'":
                if i + 1 < len(body) and body[i + 1] == "'": i += 2; continue
                instr = False
        elif c == "'": instr = True
        elif c == "(":
            if depth == 0: start = i + 1
            depth += 1
        elif c == ")":
            depth -= 1
            if depth == 0 and start is not None: yield body[start:i]; start = None
        i += 1

# ── category taxonomy: subcategory slug -> depth-1 category name ──────────────
cat_sql = open(f"{FP}/099_catalog_seed_categories.sql").read()
cats = {}
for body in re.findall(r"\bVALUES\b(.*?)\bON CONFLICT\b", cat_sql, re.S):
    for row in split_rows(body):
        f = split_top(row)
        if len(f) < 4: continue
        m = re.search(r"slug\s*=\s*'([^']+)'", f[1] or "")
        depth = next((int(unq(x)) for x in f[::-1] if isinstance(unq(x), str) and re.fullmatch(r"\d", unq(x) or "")), None)
        cats[unq(f[3])] = {"name": unq(f[2]), "parent": m.group(1) if m else None, "depth": depth}

def depth1_name(slug):
    s = slug
    while s in cats and cats[s]["depth"] not in (1, None) and cats[s]["parent"]:
        s = cats[s]["parent"]
    return cats.get(s, {}).get("name", cats.get(slug, {}).get("name", slug))

# ── XPMS classification map: depth-1 category -> (urid, dept, discipline, kind, tier, phase, xyz) ──
# URIDs are real canonical xpms_registry Team codes.
M = {
    "Bar":                        ("8000.050.00", "Hospitality", "Hospitality & F&B",  "catering",   "04 Physical", "Operate",     "Y — Variable"),
    "Kitchen":                    ("8000.060.00", "Hospitality", "Hospitality & F&B",  "catering",   "04 Physical", "Operate",     "Y — Variable"),
    "Restaurant":                 ("8000.040.00", "Hospitality", "Hospitality & F&B",  "catering",   "04 Physical", "Operate",     "Y — Variable"),
    "Catering":                   ("8000.030.00", "Hospitality", "Hospitality & F&B",  "catering",   "04 Physical", "Operate",     "Y — Variable"),
    "Green Room & Hospitality":   ("8000.080.00", "Hospitality", "Hospitality & F&B",  "catering",   "04 Physical", "Operate",     "Y — Variable"),
    "Leadership":                 ("6000.010.00", "Operations",  "Festival & Touring", "labor",      "04 Physical", "Operate",     "Z — Timeline/Phase"),
    "General Labor":              ("6000.100.00", "Operations",  "Festival & Touring", "labor",      "04 Physical", "Operate",     "Z — Timeline/Phase"),
    "Skilled Labor":              ("6000.100.00", "Operations",  "Fabrication",        "labor",      "04 Physical", "Build",       "Z — Timeline/Phase"),
    "Heavy Equipment Operators":  ("6000.020.00", "Operations",  "Construction",       "labor",      "04 Physical", "Build",       "Z — Timeline/Phase"),
    "Merchandise":                ("7000.080.00", "Experience",  "Experiential",       "equipment",  "04 Physical", "Procurement", "Y — Variable"),
    "Vendor Marketplace":         ("7000.060.00", "Experience",  "Experiential",       "equipment",  "04 Physical", "Procurement", "Y — Variable"),
    "Site Assets & Infrastructure": ("6000.020.00", "Operations","Construction",       "equipment",  "04 Physical", "Build",       "Y — Variable"),
    "Site Equipment & Tools":     ("6000.020.00", "Operations",  "Procurement",        "tool",       "04 Physical", "Procurement", "Y — Variable"),
    "Site Services":              ("5000.130.00", "Production",  "Construction",       "equipment",  "04 Physical", "Install",     "X — Constant"),
    "Site Vehicles":              ("6000.170.00", "Operations",  "Festival & Touring", "vehicle",    "04 Physical", "Procurement", "Y — Variable"),
    "Heavy Equipment":            ("4000.150.00", "Build",       "Construction",       "equipment",  "04 Physical", "Build",       "Y — Variable"),
    "Signage & Wayfinding":       ("6000.190.00", "Operations",  "Fabrication",        "equipment",  "04 Physical", "Install",     "Y — Variable"),
    "Audio":                      ("5000.030.00", "Production",  "Live Entertainment", "equipment",  "04 Physical", "Procurement", "Y — Variable"),
    "Lighting":                   ("5000.040.00", "Production",  "Live Entertainment", "equipment",  "04 Physical", "Procurement", "Y — Variable"),
    "Video":                      ("5000.050.00", "Production",  "Broadcast & Content","equipment",  "04 Physical", "Procurement", "Y — Variable"),
    "Rigging":                    ("5000.090.00", "Production",  "Live Entertainment", "equipment",  "04 Physical", "Build",       "Y — Variable"),
    "Backline":                   ("5000.100.00", "Production",  "Live Entertainment", "equipment",  "04 Physical", "Procurement", "Y — Variable"),
    "Staging":                    ("5000.110.00", "Production",  "Live Entertainment", "equipment",  "04 Physical", "Build",       "Y — Variable"),
    "Airfare":                    ("8000.090.00", "Hospitality", "Festival & Touring", "travel",     "04 Physical", "Advance",     "Y — Variable"),
    "Transportation":             ("8000.090.00", "Hospitality", "Festival & Touring", "travel",     "04 Physical", "Advance",     "Y — Variable"),
    "Rental Vehicles":            ("8000.090.00", "Hospitality", "Festival & Touring", "vehicle",    "04 Physical", "Procurement", "Y — Variable"),
    "Lodging":                    ("8000.100.00", "Hospitality", "Festival & Touring", "lodging",    "04 Physical", "Advance",     "Y — Variable"),
    "Access & Credentials":       ("6000.110.00", "Operations",  "Procurement",        "credential", "04 Physical", "Advance",     "Y — Variable"),
    "Radio & Communications":     ("9000.020.00", "Technology",  "Broadcast & Content","radio",      "04 Physical", "Procurement", "Y — Variable"),
    "Uniforms":                   ("6000.100.00", "Operations",  "Procurement",        "uniform",    "04 Physical", "Procurement", "Y — Variable"),
    "Furnishings":                ("6000.040.00", "Operations",  "Interior Design",    "equipment",  "04 Physical", "Procurement", "Y — Variable"),
    "Health & Safety":            ("6000.140.00", "Operations",  "Procurement",        "equipment",  "04 Physical", "Procurement", "Y — Variable"),
}

# ── parse items (migration 100) ───────────────────────────────────────────────
items_sql = open(f"{FP}/100_catalog_seed_items.sql").read()
COLS = ["organization_id","category_id","name","description","sku","hierarchical_sku",
        "common_name","search_aliases","options","modifiers_summary","prerequisites",
        "pricing_unit","lead_time_hours","setup_time","strike_time","crew_required",
        "power_requirements","footprint","truck_space","weather","compliance_tags",
        "sustainability_tags","unspsc_code","specifications","unit_of_measure",
        "is_custom","is_critical_path","client_visible","status","sort_order"]

def jsonb_text(v):
    if v is None: return None
    m = re.search(r"jsonb_build_object\(\s*'text'\s*,\s*'(.*)'\s*\)", v, re.S)
    return m.group(1).replace("''", "'") if m else unq(v)

records, unmapped = [], set()
for blk in re.findall(r"VALUES\s*\((.*?)\)\s*ON CONFLICT DO NOTHING;", items_sql, re.S):
    f = split_top(blk)
    row = dict(zip(COLS, f))
    m = re.search(r"slug\s*=\s*'([^']+)'", row["category_id"])
    cat = depth1_name(m.group(1) if m else "?")
    if cat not in M:
        unmapped.add(cat); continue
    urid, dept, disc, kind, tier, phase, xyz = M[cat]
    records.append({
        "code": unq(row["sku"]), "name": unq(row["name"]),
        "description": unq(row["description"]),
        "kind": kind, "urid": urid, "department": dept, "discipline": disc,
        "tier": tier, "phase": phase, "xyz": xyz,
        "specifications": jsonb_text(row["specifications"]),
        "lead_time_hours": unq(row["lead_time_hours"]), "crew": unq(row["crew_required"]),
        "source": "frozenphoenix-v6", "category": cat,
    })

if unmapped:
    raise SystemExit(f"UNMAPPED categories: {unmapped}")

# ── pricing (migration 102): name -> standard (low, high) in dollars ─────────
price_sql = open(f"{FP}/102_catalog_seed_pricing.sql").read()
std, allp = {}, {}
pat = re.compile(r"SELECT id,\s*'(\w+)'::pricing_tier,\s*'USD',\s*([\d.]+),\s*([\d.]+)\s*"
                 r"FROM catalog_items WHERE name = '((?:[^']|'')+)'", re.S)
for tier, lo, hi, nm in pat.findall(price_sql):
    nm = nm.replace("''", "'")
    allp.setdefault(nm, {})[tier] = (float(lo), float(hi))
for r in records:
    p = allp.get(r["name"], {})
    s = p.get("standard")
    r["std_low"], r["std_high"] = (s if s else (None, None))
    r["unit_cost_cents"] = int(round((s[0] + s[1]) / 2 * 100)) if s else None

# ── existing flyingbluewhale canonical items, XPMS-classified from scratch ────
FBW = [
    ("AFL-COUNTER",       "Counter / Desk — Plywood Counter w/ Graphic Skin (3'×3'2\")", "equipment", "4000.020.00", "Build",      "Fabrication",        "04 Physical", "Build"),
    ("AFL-FENCERAIL",     "Fence / Rail — Plywood Frame w/ Mesh Infill (4'×2'6\")",      "equipment", "4000.020.00", "Build",      "Fabrication",        "04 Physical", "Build"),
    ("AFL-FRAMEMODULE",   "Frame Module — Plywood Frame w/ Mesh Infill (3'×7')",         "equipment", "4000.020.00", "Build",      "Fabrication",        "04 Physical", "Build"),
    ("AFL-SHELF",         "Shelf / Display — Plywood Shelving w/ Pegboard Back (3'×6'6\")","equipment","4000.020.00", "Build",      "Fabrication",        "04 Physical", "Build"),
    ("AFL-SIGNPANEL",     "Sign Panel — Plywood Sign w/ Base (2'×6')",                   "equipment", "4000.140.00", "Build",      "Fabrication",        "04 Physical", "Build"),
    ("AFL-STAGEPLATFORM", "Stage Platform — Adjustable Height Deck (6'×1'4\")",          "equipment", "5000.110.00", "Production", "Live Entertainment", "04 Physical", "Build"),
    ("AFL-TOWERPANEL",    "Tower Panel — Plywood Framed Identity Tower (2'×8')",         "equipment", "4000.020.00", "Build",      "Fabrication",        "04 Physical", "Build"),
    ("AFL-WALLPANEL",     "Wall Panel — Plywood Frame w/ Graphic Skin (3'×7')",          "equipment", "4000.020.00", "Build",      "Fabrication",        "04 Physical", "Build"),
    ("moto-xpr7550",      "Motorola XPR7550 Two-Way Radio",                              "radio",     "9000.020.00", "Technology", "Broadcast & Content","04 Physical", "Procurement"),
    ("cred-aaa",          "All-Access Laminate",                                          "credential","6000.110.00", "Operations", "Procurement",        "04 Physical", "Advance"),
    ("legacy-ee120edb",   "All-Access Credential — Production",                           "credential","6000.110.00", "Operations", "Procurement",        "04 Physical", "Advance"),
    ("legacy-07a084d4",   "EDC LV — Mainstage Crew Catering Pack",                        "catering",  "8000.030.00", "Hospitality","Hospitality & F&B",  "04 Physical", "Operate"),
]
fbw_updates = []
existing_names = {r["name"].strip().lower() for r in records}
for code, name, kind, urid, dept, disc, tier, phase in FBW:
    fbw_updates.append({"code": code, "name": name, "kind": kind, "urid": urid,
                        "department": dept, "discipline": disc, "tier": tier,
                        "phase": phase, "xyz": "Y — Variable"})

# ── dedup combined set by normalized name (keep first) ───────────────────────
seen, deduped, dropped = set(), [], 0
for r in records:
    k = (r["name"].strip().lower())
    if k in seen: dropped += 1; continue
    seen.add(k); deduped.append(r)
records = deduped

# ── emit SQL ──────────────────────────────────────────────────────────────────
def q(v):
    if v is None: return "NULL"
    if isinstance(v, (int, float)): return str(v)
    return "'" + str(v).replace("'", "''") + "'"

schema_sql = f"""-- XPMS Master Catalog — schema + cleanup (migration 1 of 2)
-- Adds XPMS classification axes to master_catalog_items, extends catalog_kind
-- with 'labor', and soft-deletes the duplicated external_example Casa rows.

ALTER TYPE public.catalog_kind ADD VALUE IF NOT EXISTS 'labor';

ALTER TABLE public.master_catalog_items
  ADD COLUMN IF NOT EXISTS urid            text,
  ADD COLUMN IF NOT EXISTS xpms_department text,
  ADD COLUMN IF NOT EXISTS discipline      text,
  ADD COLUMN IF NOT EXISTS default_tier    text,
  ADD COLUMN IF NOT EXISTS default_phase   text,
  ADD COLUMN IF NOT EXISTS xyz             text;

COMMENT ON COLUMN public.master_catalog_items.urid IS
  'Canonical XPMS URID DEPT.TEAM.SECTION (maps to public.xpms_registry).';

CREATE INDEX IF NOT EXISTS idx_master_catalog_items_urid
  ON public.master_catalog_items (urid) WHERE deleted_at IS NULL;
CREATE INDEX IF NOT EXISTS idx_master_catalog_items_department
  ON public.master_catalog_items (xpms_department) WHERE deleted_at IS NULL;

-- Purge duplicated demo seed rows (reversible soft-delete).
UPDATE public.master_catalog_items
   SET deleted_at = now(), active = false
 WHERE org_id = {q(ORG)} AND scope = 'external_example' AND deleted_at IS NULL;
"""

data_lines = ["-- XPMS Master Catalog — data load (migration 2 of 2)",
              "-- Inserts the combined, deduped, XPMS-classified catalog (canonical).", ""]
for r in records:
    data_lines.append(
        "INSERT INTO public.master_catalog_items "
        "(org_id, kind, code, name, description, unit_cost_cents, currency, active, scope, "
        "urid, xpms_department, discipline, default_tier, default_phase, xyz) VALUES ("
        f"{q(ORG)}, {q(r['kind'])}::catalog_kind, {q(r['code'])}, {q(r['name'])}, {q(r['description'])}, "
        f"{q(r['unit_cost_cents'])}, 'USD', true, 'canonical', "
        f"{q(r['urid'])}, {q(r['department'])}, {q(r['discipline'])}, {q(r['tier'])}, {q(r['phase'])}, {q(r['xyz'])}) "
        "ON CONFLICT (org_id, code) DO UPDATE SET "
        "kind=EXCLUDED.kind, name=EXCLUDED.name, description=EXCLUDED.description, "
        "unit_cost_cents=EXCLUDED.unit_cost_cents, scope='canonical', active=true, deleted_at=NULL, "
        "urid=EXCLUDED.urid, xpms_department=EXCLUDED.xpms_department, discipline=EXCLUDED.discipline, "
        "default_tier=EXCLUDED.default_tier, default_phase=EXCLUDED.default_phase, xyz=EXCLUDED.xyz, "
        "updated_at=now();")
data_lines.append("\n-- Classify existing flyingbluewhale canonical items into XPMS.")
for u in fbw_updates:
    data_lines.append(
        f"UPDATE public.master_catalog_items SET "
        f"urid={q(u['urid'])}, xpms_department={q(u['department'])}, discipline={q(u['discipline'])}, "
        f"default_tier={q(u['tier'])}, default_phase={q(u['phase'])}, xyz={q(u['xyz'])}, updated_at=now() "
        f"WHERE org_id={q(ORG)} AND code={q(u['code'])} AND deleted_at IS NULL;")
data_sql = "\n".join(data_lines) + "\n"

open(f"{ROOT}/supabase/migrations/20260608120000_xpms_master_catalog_schema.sql", "w").write(schema_sql)
open(f"{ROOT}/supabase/migrations/20260608120100_xpms_master_catalog_data.sql", "w").write(data_sql)
json.dump({"records": records, "fbw": fbw_updates}, open(f"{ROOT}/scripts/.xpms-catalog.json", "w"), indent=1)

# ── xlsx (matches the applied catalog) ────────────────────────────────────────
wb = openpyxl.Workbook(); RED = "B91C1C"
hf, hfont = PatternFill("solid", fgColor=RED), Font(bold=True, color="FFFFFF")
def mk(title, headers, rows, widths, wrap_cols=()):
    ws = wb.create_sheet(title); ws.append(headers)
    for c in range(1, len(headers) + 1):
        ws.cell(1, c).fill = hf; ws.cell(1, c).font = hfont
        ws.column_dimensions[get_column_letter(c)].width = widths[c - 1]
    for row in rows:
        ws.append(row)
    for c in wrap_cols:
        for rr in range(2, len(rows) + 2):
            ws.cell(rr, c).alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"; ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows)+1}"
    return ws
wb.remove(wb.active)

allrows = records + [{"code": u["code"], "name": u["name"], "kind": u["kind"], "urid": u["urid"],
                      "department": u["department"], "discipline": u["discipline"], "tier": u["tier"],
                      "phase": u["phase"], "xyz": u["xyz"], "unit_cost_cents": None, "std_low": None,
                      "std_high": None, "specifications": None, "lead_time_hours": None, "crew": None,
                      "source": "flyingbluewhale"} for u in fbw_updates]

from collections import Counter
by_dept = Counter(r["department"] for r in allrows)
by_kind = Counter(r["kind"] for r in allrows)
ov = [["By XPMS Department", ""]] + [[d, n] for d, n in sorted(by_dept.items(), key=lambda x: -x[1])]
ov += [["", ""], ["By catalog_kind", ""]] + [[k, n] for k, n in sorted(by_kind.items(), key=lambda x: -x[1])]
ov += [["", ""], ["TOTAL ITEMS", len(allrows)], ["From FrozenPhoenix v6", len(records)],
       ["From flyingbluewhale", len(fbw_updates)], ["Deduped (dropped)", dropped]]
mk("Overview", ["Group", "Count"], ov, [30, 10])

headers = ["Code", "Name", "Kind", "URID", "Department", "Discipline", "Default Tier",
           "Default Phase", "XYZ", "Unit Cost USD", "Std $ Low", "Std $ High", "Source",
           "Description", "Specifications", "Lead Time (hrs)", "Crew"]
widths = [20, 46, 12, 12, 13, 20, 13, 13, 18, 13, 10, 10, 16, 50, 46, 14, 22]
rows = []
for r in allrows:
    uc = r.get("unit_cost_cents")
    rows.append([r["code"], r["name"], r["kind"], r["urid"], r["department"], r["discipline"],
                 r["tier"], r["phase"], r["xyz"], (uc / 100 if uc is not None else None),
                 r.get("std_low"), r.get("std_high"), r["source"], r.get("description"),
                 r.get("specifications"), r.get("lead_time_hours"), r.get("crew")])
ws = mk("Master XPMS Catalog", headers, rows, widths, wrap_cols=(14, 15))
ws.column_dimensions["J"].number_format = "#,##0.00"
wb.save(f"{ROOT}/xpms-master-catalog.xlsx")

print(f"FrozenPhoenix mapped: {len(records)} (deduped {dropped}) · fbw: {len(fbw_updates)} · total: {len(allrows)}")
print(f"By department: " + ", ".join(f'{d}:{n}' for d, n in sorted(by_dept.items(), key=lambda x:-x[1])))
print(f"By kind: " + ", ".join(f'{k}:{n}' for k, n in sorted(by_kind.items(), key=lambda x:-x[1])))
print("Wrote 2 migration files + xpms-master-catalog.xlsx")
