#!/usr/bin/env python3
"""Export the Universal Advance Seed Catalog v6.0 (351 items) from the
FrozenPhoenix SQL seed migrations into a single .xlsx.

Sources (read-only):
  099_catalog_seed_categories.sql  -> category taxonomy (slug -> name/path)
  100_catalog_seed_items.sql       -> 351 catalog_items rows
  102_catalog_seed_pricing.sql     -> 3-tier USD pricing (low/high) per item
"""
import re
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

FP = "/Users/julianclarkson/Documents/FrozenPhoenix/supabase/migrations"
OUT = "/Users/julianclarkson/Documents/flyingbluewhale/advancing-seed-catalog-351.xlsx"

COLS = ["organization_id","category_id","name","description","sku","hierarchical_sku",
        "common_name","search_aliases","options","modifiers_summary","prerequisites",
        "pricing_unit","lead_time_hours","setup_time","strike_time","crew_required",
        "power_requirements","footprint","truck_space","weather","compliance_tags",
        "sustainability_tags","unspsc_code","specifications","unit_of_measure",
        "is_custom","is_critical_path","client_visible","status","sort_order"]


def split_top(s):
    """Split a SQL VALUES body on top-level commas (respect quotes/parens/brackets)."""
    out, buf, depth, instr = [], [], 0, False
    i = 0
    while i < len(s):
        c = s[i]
        if instr:
            if c == "'":
                if i + 1 < len(s) and s[i + 1] == "'":  # '' escape
                    buf.append("''"); i += 2; continue
                instr = False
            buf.append(c)
        else:
            if c == "'":
                instr = True; buf.append(c)
            elif c in "([":
                depth += 1; buf.append(c)
            elif c in ")]":
                depth -= 1; buf.append(c)
            elif c == "," and depth == 0:
                out.append("".join(buf).strip()); buf = []
            else:
                buf.append(c)
        i += 1
    if buf:
        out.append("".join(buf).strip())
    return out


def unq(v):
    v = v.strip()
    if v.upper() == "NULL":
        return None
    # strip a trailing ::type cast
    v = re.sub(r"::\w+(\s*\[\])?$", "", v).strip()
    if v.startswith("'") and v.endswith("'"):
        return v[1:-1].replace("''", "'")
    if v.upper() in ("TRUE", "FALSE"):
        return v.upper() == "TRUE"
    if re.fullmatch(r"-?\d+", v):
        return int(v)
    return v


def parse_array(v):
    if v is None or not v.strip().upper().startswith("ARRAY"):
        return None
    inner = v.strip()[v.strip().index("[") + 1: v.strip().rindex("]")]
    if not inner.strip():
        return ""
    return "; ".join(unq(x) for x in split_top(inner) if unq(x) is not None)


def parse_jsonb_text(v):
    if v is None:
        return None
    m = re.search(r"jsonb_build_object\(\s*'text'\s*,\s*'(.*)'\s*\)", v, re.S)
    if m:
        return m.group(1).replace("''", "'")
    return unq(v)


def slug_from_subquery(v):
    if v is None:
        return None
    m = re.search(r"slug\s*=\s*'([^']+)'", v)
    return m.group(1) if m else unq(v)


def split_rows(body):
    """Yield each top-level (...) group from a multi-row VALUES body."""
    depth, instr, start = 0, False, None
    i = 0
    while i < len(body):
        c = body[i]
        if instr:
            if c == "'":
                if i + 1 < len(body) and body[i + 1] == "'":
                    i += 2; continue
                instr = False
        elif c == "'":
            instr = True
        elif c == "(":
            if depth == 0:
                start = i + 1
            depth += 1
        elif c == ")":
            depth -= 1
            if depth == 0 and start is not None:
                yield body[start:i]
                start = None
        i += 1


# ── categories: slug -> name (cols: org, parent, name=field2, slug=field3) ────
cat_name = {}
with open(f"{FP}/099_catalog_seed_categories.sql") as f:
    cat_sql = f.read()
for body in re.findall(r"\bVALUES\b(.*?)\bON CONFLICT\b", cat_sql, re.S):
    for row in split_rows(body):
        fields = split_top(row)
        if len(fields) < 4:
            continue
        name, slug = unq(fields[2]), unq(fields[3])
        if isinstance(slug, str) and slug and slug not in cat_name:
            cat_name[slug] = name or slug

# ── items ─────────────────────────────────────────────────────────────────────
with open(f"{FP}/100_catalog_seed_items.sql") as f:
    items_sql = f.read()
blocks = re.findall(r"VALUES\s*\((.*?)\)\s*ON CONFLICT DO NOTHING;", items_sql, re.S)
items = []
for blk in blocks:
    fields = split_top(blk)
    if len(fields) != len(COLS):
        raise SystemExit(f"field count {len(fields)} != {len(COLS)} in block:\n{blk[:200]}")
    row = dict(zip(COLS, fields))
    rec = {
        "category_slug": slug_from_subquery(row["category_id"]),
        "name": unq(row["name"]),
        "common_name": unq(row["common_name"]),
        "sku": unq(row["sku"]),
        "description": unq(row["description"]),
        "search_aliases": parse_array(row["search_aliases"]),
        "options": parse_array(row["options"]),
        "modifiers_summary": unq(row["modifiers_summary"]),
        "prerequisites": unq(row["prerequisites"]),
        "pricing_unit": unq(row["pricing_unit"]),
        "lead_time_hours": unq(row["lead_time_hours"]),
        "setup_time": unq(row["setup_time"]),
        "strike_time": unq(row["strike_time"]),
        "crew_required": unq(row["crew_required"]),
        "power_requirements": unq(row["power_requirements"]),
        "footprint": unq(row["footprint"]),
        "truck_space": unq(row["truck_space"]),
        "weather": unq(row["weather"]),
        "compliance_tags": parse_array(row["compliance_tags"]),
        "sustainability_tags": parse_array(row["sustainability_tags"]),
        "unspsc_code": unq(row["unspsc_code"]),
        "specifications": parse_jsonb_text(row["specifications"]),
        "unit_of_measure": unq(row["unit_of_measure"]),
        "is_critical_path": unq(row["is_critical_path"]),
        "client_visible": unq(row["client_visible"]),
        "status": unq(row["status"]),
        "sort_order": unq(row["sort_order"]),
    }
    rec["category"] = cat_name.get(rec["category_slug"], rec["category_slug"])
    items.append(rec)

# ── pricing: item name -> {tier: (low, high)} ────────────────────────────────
with open(f"{FP}/102_catalog_seed_pricing.sql") as f:
    price_sql = f.read()
prices = {}
pat = re.compile(
    r"SELECT id,\s*'(\w+)'::pricing_tier,\s*'USD',\s*([\d.]+),\s*([\d.]+)\s*"
    r"FROM catalog_items WHERE name = '((?:[^']|'')+)'",
    re.S,
)
for tier, lo, hi, nm in pat.findall(price_sql):
    nm = nm.replace("''", "'")
    prices.setdefault(nm, {})[tier] = (float(lo), float(hi))

# ── workbook ─────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
RED = "B91C1C"
hfill = PatternFill("solid", fgColor=RED)
hfont = Font(bold=True, color="FFFFFF")


def sheet(title, headers, rows, widths):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        ws.cell(1, c).fill = hfill
        ws.cell(1, c).font = hfont
        ws.column_dimensions[get_column_letter(c)].width = widths[c - 1]
    for r in rows:
        ws.append(r)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"
    return ws


wb.remove(wb.active)

# Overview
collections = {}
for it in items:
    coll = (it["sku"].split("-")[0] if it["sku"] else "?")
    collections[coll] = collections.get(coll, 0) + 1
ov_rows = [[k, v] for k, v in sorted(collections.items(), key=lambda x: -x[1])]
sheet("Overview",
      ["SKU Prefix (Collection)", "Items"],
      ov_rows + [["TOTAL", len(items)]],
      [28, 10])

# Items
item_headers = ["SKU", "Name", "Common Name", "Category", "Category Slug", "Description",
                "Specifications", "Search Aliases", "Options", "Modifiers", "Prerequisites",
                "Pricing Unit", "UoM", "Lead Time (hrs)", "Setup", "Strike", "Crew",
                "Power", "Footprint", "Truck Space", "Weather", "Sustainability Tags",
                "Compliance Tags", "UNSPSC", "Critical Path", "Client Visible", "Status",
                "Basic $ Low", "Basic $ High", "Std $ Low", "Std $ High", "Prem $ Low", "Prem $ High"]
item_widths = [20, 40, 24, 26, 20, 60, 50, 40, 36, 26, 30, 16, 10, 14, 18, 18, 18,
               16, 18, 28, 12, 24, 18, 12, 12, 12, 9, 11, 11, 11, 11, 11, 11]
item_rows = []
for it in items:
    p = prices.get(it["name"], {})
    b, s, pr = p.get("basic", (None, None)), p.get("standard", (None, None)), p.get("premium", (None, None))
    item_rows.append([
        it["sku"], it["name"], it["common_name"], it["category"], it["category_slug"],
        it["description"], it["specifications"], it["search_aliases"], it["options"],
        it["modifiers_summary"], it["prerequisites"], it["pricing_unit"], it["unit_of_measure"],
        it["lead_time_hours"], it["setup_time"], it["strike_time"], it["crew_required"],
        it["power_requirements"], it["footprint"], it["truck_space"], it["weather"],
        it["sustainability_tags"], it["compliance_tags"], it["unspsc_code"],
        it["is_critical_path"], it["client_visible"], it["status"],
        b[0], b[1], s[0], s[1], pr[0], pr[1],
    ])
sheet("Advancing Catalog (351)", item_headers, item_rows, item_widths)

wb.save(OUT)
priced = sum(1 for it in items if it["name"] in prices)
print(f"Wrote {OUT}")
print(f"  items={len(items)}  categories={len(cat_name)}  priced={priced}")
print(f"  collections: " + ", ".join(f'{k}:{v}' for k, v in sorted(collections.items(), key=lambda x: -x[1])))
